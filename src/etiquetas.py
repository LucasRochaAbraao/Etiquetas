# coding=utf-8
import random
import logging
import openpyxl
import os.path
from datetime import datetime

# TODO:
# - Modo manual e gerar automaticamente [300] de cada.
# Se for gerar 300 de cada, apartir de qual número?
# o problema é rodar o programa sem ser para produção
# e o log estragar a sequência.
# Uma opção seria rodar com -d sim por padrão e caso
# seja para produção passar um parâmetro. e.g.: -p 250
# - Execute the program with a GUI

### PATHS ###
# inicializar diretórios de arquivos e logs.
# first returns __file__, then climbs to it's parent's parent
PROJECT_PATH = os.path.abspath(__file__ + "/../..") # returns /etiquetas
#FILEPATH    = os.path.dirname(__file__) returns directory name
LOG_PATH      = PROJECT_PATH + '/log'
EMAIL        = PROJECT_PATH + '/email'
SRC_FILE_BGN   = PROJECT_PATH + '/templates/template-bgn.xlsx'
END_FILE_BGN   = 'etiquetas-bgn.xlsx'
SRC_FILE_AC  = PROJECT_PATH + '/templates/template-ac.xlsx'
END_FILE_AC  = 'etiquetas-ac.xlsx'



class Etiquetas():
    ############### init setup #################
    # define os caminhos padrões e inicializa o "workbook" e "worksheet"
    def __init__(self, inicio, fim, modelo_comum, debug):
        ### CONFIG LOGGER ###
        LOG_FORMAT = "%(asctime)s [%(levelname)s]:\t%(message)s"
        logging.basicConfig(filename=f'{LOG_PATH}/logs', format=LOG_FORMAT)
        self.logger = logging.getLogger(__name__)
        self.logger.setLevel(logging.INFO)
        # inicializar parâmetros
        self.inicio = inicio
        self.fim = fim + 1
        self.soma = self.fim - self.inicio
        self.modelo_comum = modelo_comum
        self.debug = debug
        if self.debug:
            self.logger.setLevel(logging.DEBUG)

        if self.modelo_comum:
            self.WB = openpyxl.load_workbook(SRC_FILE_BGN)
            self.WS = self.WB.active
            self.logger.debug("Modelo Comum Selecionado")
        else:
            self.WB = openpyxl.load_workbook(SRC_FILE_AC)
            self.WS = self.WB.active
            self.logger.debug("Modelo AC Selecionado")

        #self.gerar(self.inicio, self.fim)

    def get_alpha(self):
        # retorna as letras em formato de lista. OBS: Sem as letras I L O
        # para não gerar confusão na leitura da etiqueta (I,l,1,0,o...?)
        return list('abcdefghjkmnpqrstuvwxyz')
    def get_numerico(self):
        # retorn 8 números em formato de número. OBS: Sem os números 0 e 1.
        return list(range(2,10))

    # criar SSIDs
    def get_ssid_comum(self, incremento):
        return "QUICK/+FIBRA_" + str(incremento)
    def get_ssid_bgn(self, incremento):
        return "QUICK/+FIBRA_" + str(incremento) + '-2.4GHz'
    def get_ssid_ac(self, incremento):
        return "QUICK/+FIBRA_" + str(incremento) + '-5GHz'

    def gerar_alpha(self, strings):
        # recebe a lista de letras e retorna uma nova lista com 4 letras aleatórias
        rand_alpha = []
        for _ in range(4):
            rand_alpha.append(random.choice(strings))
        return rand_alpha

    def gerar_numerico(self, numericos):
        # recebe a lista de números e retorna uma nova lista com 4 números aleatórios
        rand_numerico = []
        for _ in range(4):
            rand_numerico.append(random.choice(numericos))
        return rand_numerico

    def gerar_senha(self, strings, numericos):
        # Recebe a lista completa de letras e números. Passa cada lista para gerar
        # 4 dígitos aleatórios de cada. Junta cada lista em uma varíavel (agora com 8 itens)
        # e retorna essa lista embaralhada. 
        alpha = self.gerar_alpha(strings)
        numerico = self.gerar_numerico(numericos)
        rand_senha = alpha + [str(num) for num in numerico]
        random.shuffle(rand_senha)
        return rand_senha

    def gerar(self):
        # Primeiro confere se o início é menor ou igual ao fim.
        # Depois gera as listas completas de letras e números.
        if self.fim <= self.inicio:
            msg = "Fim menor do que início! Alcance não válido!"
            self.logger.critical(msg)
            raise Exception(msg)
        lista_alpha = self.get_alpha()
        lista_numer = self.get_numerico()

        # - inicializar a lista dos números incrementais. Deixei fora da
        # função que gera os SSIDs pq no futuro fica fácil isso ser um input
        # do usuário ou ler de um arquivo e passar pronto pra função.
        # - Depois apenas inicializo as listas que guardarão os SSIDs.
        ordem = [i for i in range(self.inicio, self.fim)]
        ssid_comum    = [] # ssid comum
        ssid_bgn  = [] # ssid b/g/n
        ssid_ac = [] # ssid ac
        senha   = [] # 4 letras e 4 números aleatórios.

        for i in range(self.fim - self.inicio): # quantidade gerada
            # (fim - início) gera a quantidade de ssids que a função vai criar,
            # ou seja, quantas vezes vai passar por esse for loop. Ao gerar
            # os ssids, passo a lista "ordem" para usar os números passados
            # como início e fim específicos.
            temp_ssid_comum  = self.get_ssid_comum(ordem[i])
            ssid_comum.append(temp_ssid_comum)

            temp_ssid_bgn = self.get_ssid_bgn(ordem[i])
            ssid_bgn.append(temp_ssid_bgn)

            temp_ssid_ac = self.get_ssid_ac(ordem[i])
            ssid_ac.append(temp_ssid_ac)

            # senhas
            temp_senha = self.gerar_senha(lista_alpha, lista_numer)
            senha.append(temp_senha)

        # aqui retorno todos os SSIDs e a senha, quem receber essas
        # informações decide o que fazer com elas. (a/b/n vs ac)
        self.escrever_dados(ssid_comum, ssid_bgn, ssid_ac, senha)
        if self.debug:
            print('\t\t +++ Entrando em modo debug +++')
            self.debuggar(ssid_comum, ssid_bgn, ssid_ac, senha)

    def escrever_dados(self, ssid_comum, ssid_bgn, ssid_ac, senha):
            ############# GRAVAR DADOS ###############
            # primeiro especifica a coluna do arquivo a ser gravada e
            # depois coloca os valores na celula. O enumerate gera
            # um índice de acordo com o tamanho das listas, usado
            # para identificar qual número de célula para gravar (i).
            # o "+2" é para pular as 2 colunas iniciais (título e célula em branco)

            if self.modelo_comum:  # se o modelo for comum, salva o ssid + senha
                # ssid comum
                column = 1
                for row, value in enumerate(ssid_comum):
                    self.WS.cell(column=column, row=row + 2, value=value)
                # senha
                column = 2
                for row, value in enumerate(senha):
                    self.WS.cell(column=column, row=row + 2, value=''.join(value))

            else:     # caso contrário, salva os SSIDs + senha do outro modelo
                # 2.4GHz    
                column = 1
                for row, value in enumerate(ssid_bgn):
                    self.WS.cell(column=column, row=row + 2, value=value)
                # 5GHz
                column = 2
                for row, value in enumerate(ssid_ac):
                    self.WS.cell(column=column, row=row + 2, value=value)
                # senha
                column = 3
                for row, value in enumerate(senha):
                    self.WS.cell(column=column, row=row + 2, value=''.join(value))
            #salvar()

    def debuggar(self, ssid_comum, ssid_bgn, ssid_ac, senha):
        print("================")
        if self.modelo_comum:
            for onu, pw in zip(ssid_comum, senha):
                print(onu)
                print(''.join(pw)) # juntar os caracteres numa string só, e não tudo separado.
                print("================")
        else:
            for modelo_bgn, modelo_ac, pw in zip(ssid_bgn, ssid_ac, senha):
                print(modelo_bgn)
                print(modelo_ac)
                print(''.join(pw))
                print("================")

    #relativo_dir  = datetime.now().strftime('logs/OLTs/' + olt + '-%Y_%m_%d-%H_%M_%S.log')

    def salvar(self):
        data = datetime.now().strftime("%d-%m-%Y")
        hora = datetime.now().strftime("%H:%M:%S")
        if self.modelo_comum:
            self.WB.save(f'{EMAIL}/{END_FILE_BGN}')
            #self.WB.save(f'{LOG_PATH}/{data}_{hora}_{str(self.soma)}_{END_FILE_BGN}')
            if os.path.isfile(f'{EMAIL}/{END_FILE_BGN}'):
                self.logger.info(f"{self.soma} ({self.inicio}-{self.fim-1}) etiquetas BGN arquivadas em 'email/{END_FILE_BGN}'.")
            else:
                msg = "Arquivo do modelo B/G/N não foi gerado ou salvo!"
                self.logger.critical(msg)
                raise Exception(msg)
        else:
            self.WB.save(f'{EMAIL}/{END_FILE_AC}')
            #self.WB.save(f'{LOG_PATH}/{data}_{hora}_{str(self.soma)}_{END_FILE_AC}')
            if os.path.isfile(f'{EMAIL}/{END_FILE_AC}'):
                self.logger.info(f"{self.soma} etiquetas AC arquivadas em 'email/{END_FILE_AC}'.")
            else:
                msg = "Arquivo do modelo AC não foi gerado ou salvo!"
                self.logger.critical(msg)
                raise Exception(msg)
        print("Etiquetas geradas!")
